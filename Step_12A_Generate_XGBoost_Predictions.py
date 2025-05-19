"""
Step 12A: Generate XGBoost Predictions

This script generates predictions for the next month's returns using the trained XGBoost models.
It processes all factor-window combinations in parallel, making predictions on the most recent
available data for each factor. The results are saved in both HDF5 and Excel formats for further
analysis and portfolio construction.

Key Features:
- Parallel processing for efficient prediction generation
- Handles missing data and models gracefully
- Produces both machine-readable (HDF5) and human-readable (Excel) outputs
- Includes comprehensive logging and error handling

----------------------------------------------------------------------------------------------------
INPUT FILES:
- S8_Feature_Sets.h5
  - Path: ./output/S8_Feature_Sets.h5
  - Description: Feature sets for each factor and window, containing the most recent data points
                for prediction.

- S11A_XGBoost_Models/
  - Path: ./output/S11A_XGBoost_Models/
  - Description: Trained XGBoost models for each factor and window combination.
  - Format: One .joblib file per model, named as 'window_<window_id>_factor_<factor_id>.joblib'

OUTPUT FILES:
1. S12A_XGBoost_Predictions.h5
   - Path: ./output/S12A_XGBoost_Predictions.h5
   - Description: Predictions for next month's returns for all factors and windows.
   - Format: HDF5 with structure: /window_<window_id>/factor_<factor_id>/predictions
   - Use: For programmatic access to prediction data

2. S12A_XGBoost_Predictions_Matrix.xlsx
   - Path: ./output/S12A_XGBoost_Predictions_Matrix.xlsx
   - Format: Excel workbook with two sheets:
     - 'Predictions': Matrix with factors as rows and windows as columns
     - 'Summary': Key statistics about the predictions
   - Features:
     - Formatted for easy reading
     - Filterable and sortable columns
     - Color-coded headers

3. S12A_Prediction_Log.log
   - Path: ./output/S12A_Prediction_Log.log
   - Description: Detailed log of the prediction process, including any warnings or errors.
   - Use: For debugging and monitoring the prediction process

----------------------------------------------------------------------------------------------------
USAGE:
    python Step_12A_Generate_XGBoost_Predictions.py

NOTES:
- The script will automatically use all available CPU cores minus 2 for parallel processing.
- Missing models or features will be logged but won't stop the execution.
- The Excel output is formatted for easy manual inspection and analysis.
"""

import os
import h5py
import joblib
import numpy as np
import pandas as pd
from pathlib import Path
import logging
from tqdm import tqdm
import multiprocessing
from typing import Dict, List, Tuple, Optional, Set
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Suppress warnings
warnings.filterwarnings('ignore')
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('output/S12A_Prediction_Log.log'),
        logging.StreamHandler()
    ]
)

# Constants
OUTPUT_DIR = Path("output")
MODELS_DIR = OUTPUT_DIR / "S11A_XGBoost_Models"
FEATURES_PATH = OUTPUT_DIR / "S8_Feature_Sets.h5"
PREDICTIONS_PATH = OUTPUT_DIR / "S12A_XGBoost_Predictions.h5"
EXCEL_OUTPUT_PATH = OUTPUT_DIR / "S12A_XGBoost_Predictions_Matrix.xlsx"
NUM_CORES = max(1, multiprocessing.cpu_count() - 2)  # Leave some cores free


def load_model(model_path: Path) -> Optional[object]:
    """Load a trained XGBoost model from disk."""
    try:
        model = joblib.load(model_path)
        return model
    except Exception as e:
        logging.error(f"Error loading model {model_path}: {str(e)}")
        return None


def get_latest_features(h5_file: h5py.File, window_id: str, factor_id: str) -> Optional[np.ndarray]:
    """Extract the most recent feature vector for prediction."""
    try:
        # Path to the test features (most recent data)
        test_path = f"window_{window_id}/factor_{factor_id}/test"
        X_test = h5_file[f"{test_path}/X/data"][:]
        
        # Get the most recent feature vector (last row)
        if len(X_test) > 0:
            return X_test[-1:]
        return None
    except Exception as e:
        logging.error(f"Error loading features for window {window_id}, factor {factor_id}: {str(e)}")
        return None


def process_prediction_task(args):
    """Process a single model prediction task."""
    model_path, h5_path, predictions_path, lock = args
    
    # Convert string path to Path object if needed
    if isinstance(model_path, str):
        model_path = Path(model_path)
    
    # Extract factor_id from the model filename
    # Model filenames are like: factor_10Yr Bond 12_CS.joblib
    model_stem = model_path.stem
    if model_stem.startswith('factor_'):
        # Extract the factor name by removing the 'factor_' prefix
        factor_id = model_stem[7:]  # Skip the 'factor_' prefix
    else:
        logging.error(f"Invalid model filename format: {model_stem}")
        return None
    
    # Get window_id from the parent directory name
    window_id = model_path.parent.name.split('_')[1]
    
    # Filter out 3m, 12m, and 60m factors
    if any(suffix in factor_id for suffix in ["_3m", "_12m", "_60m"]):
        return None
    
    # Load model
    try:
        model = joblib.load(model_path)
    except Exception as e:
        logging.error(f"Error loading model {model_path}: {str(e)}")
        return None
    
    # Load features - FIXED PATH
    X_pred = None
    try:
        with h5py.File(h5_path, 'r') as h5f:
            # FIXED: The correct path to prediction features
            features_key = f"feature_sets/window_{window_id}/prediction/{factor_id}/X/data"
            
            if features_key in h5f:
                X_pred_data = h5f[features_key][:]
                
                # Get column names
                columns_key = f"feature_sets/window_{window_id}/prediction/{factor_id}/X/columns"
                if columns_key in h5f:
                    columns = [col.decode('utf-8') if isinstance(col, bytes) else col 
                              for col in h5f[columns_key][:]]
                    
                    # Create DataFrame with proper column names
                    X_pred = pd.DataFrame(X_pred_data, columns=columns)
                else:
                    # Use generic column names if specific ones aren't available
                    X_pred = pd.DataFrame(X_pred_data)
                    logging.warning(f"No column names found for {window_id}, {factor_id}. Using generic names.")
            else:
                logging.warning(f"No features found at {features_key}")
                return None
    except Exception as e:
        logging.error(f"Error loading features for window {window_id}, factor {factor_id}: {str(e)}")
        return None
    
    if X_pred is None or X_pred.empty:
        logging.warning(f"Empty features for window {window_id}, factor {factor_id}")
        return None
        
    # Generate prediction
    try:
        pred = model.predict(X_pred)
        
        # Check if pred is valid
        if pred is None or len(pred) == 0:
            logging.warning(f"Empty prediction for window {window_id}, factor {factor_id}")
            return None
            
        # Save prediction to HDF5 with lock to avoid concurrent writes
        with lock:
            with h5py.File(predictions_path, 'a') as h5f:
                # Create window group if it doesn't exist
                if f"window_{window_id}" not in h5f:
                    h5f.create_group(f"window_{window_id}")
                
                window_group = h5f[f"window_{window_id}"]
                
                # Create factor group if it doesn't exist
                if factor_id not in window_group:
                    window_group.create_group(factor_id)
                
                factor_group = window_group[factor_id]
                
                # Save prediction
                if 'prediction' in factor_group:
                    del factor_group['prediction']
                factor_group.create_dataset('prediction', data=pred)
                
        return (window_id, factor_id, float(pred[0]))  # Return the prediction value too for debugging
    except Exception as e:
        logging.error(f"Error generating prediction for window {window_id}, factor {factor_id}: {str(e)}")
        return None


def main():
    # Create output directory if it doesn't exist
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Clear existing predictions file
    if PREDICTIONS_PATH.exists():
        os.remove(PREDICTIONS_PATH)
    
    # Get list of all model files
    model_files = list(MODELS_DIR.glob("**/*.joblib"))
    if not model_files:
        logging.error(f"No model files found in {MODELS_DIR}")
        return
    
    logging.info(f"Found {len(model_files)} model files to process.")
    
    # Create multiprocessing lock for file access
    lock = multiprocessing.Manager().Lock()
    
    # Create tasks for parallel processing
    tasks = []
    for model_path in model_files:
        tasks.append((model_path, str(FEATURES_PATH), str(PREDICTIONS_PATH), lock))
    
    logging.info(f"Created {len(tasks)} tasks for parallel processing.")
    
    # Process predictions in parallel
    predictions = {}
    with multiprocessing.Pool(processes=NUM_CORES) as pool:
        # Use tqdm for progress bar
        results = list(tqdm(
            pool.imap(process_prediction_task, tasks),
            total=len(tasks),
            desc="Generating predictions",
            unit="model"
        ))
        
        # Organize predictions by window and factor
        for result in results:
            if result is None:
                continue
                
            window_id, factor_id, pred_value = result
            if window_id not in predictions:
                predictions[window_id] = {}
            predictions[window_id][factor_id] = pred_value
    
    # Save predictions to HDF5
    with h5py.File(PREDICTIONS_PATH, 'w') as h5f:
        for window_id, factors in predictions.items():
            window_group = h5f.create_group(f"window_{window_id}")
            for factor_id, pred in factors.items():
                factor_group = window_group.create_group(factor_id)
                factor_group.create_dataset('prediction', data=np.array([pred]))
    
    logging.info(f"Predictions saved to {PREDICTIONS_PATH}")
    
    # Save predictions to Excel in matrix format
    save_predictions_to_excel(predictions, EXCEL_OUTPUT_PATH)
    logging.info(f"Predictions matrix saved to {EXCEL_OUTPUT_PATH}")


def save_predictions_to_excel(predictions: Dict[str, Dict[str, float]], output_path: Path) -> None:
    """
    Save predictions to an Excel file with factors as columns and dates as rows.
    
    Args:
        predictions: Nested dictionary of predictions {window_id: {factor_id: prediction}}
        output_path: Path to save the Excel file
    """
    # Load window schedule to get actual dates
    try:
        window_schedule = pd.read_excel('output/S4_Window_Schedule.xlsx')
        window_dates = {}
        for _, row in window_schedule.iterrows():
            window_id = str(row['Window_ID'])
            # Use the prediction date of the window
            prediction_date = row['Prediction_Date']
            window_dates[window_id] = prediction_date
    except Exception as e:
        logging.warning(f"Could not load window schedule: {e}. Using window IDs instead.")
        window_dates = {}
    
    # Get unique factors and window IDs
    all_factors = set()
    window_ids = set()
    
    # First pass: collect all unique factors and window IDs
    for window_id, factors in predictions.items():
        window_ids.add(window_id)
        for factor_id in factors.keys():
            all_factors.add(factor_id)
    
    # Convert to sorted lists
    all_factors = sorted(all_factors)
    window_ids = sorted(window_ids, key=lambda x: int(x))  # Sort windows numerically
    
    # Create a DataFrame with windows as rows and factors as columns
    df = pd.DataFrame(index=window_ids, columns=all_factors, dtype=float)
    
    # Fill in the predictions
    for window_id, factors in predictions.items():
        for factor_id, pred in factors.items():
            if factor_id in all_factors:
                df.at[window_id, factor_id] = pred
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write predictions sheet
        df.to_excel(writer, sheet_name='Predictions', float_format='%.2f')
        
        # Format the sheet
        worksheet = writer.sheets['Predictions']
        
        # Formatting
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Format header row (factor names)
        for col in range(2, len(all_factors) + 2):  # Start from column B (2)
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format window IDs column and replace with dates
        for row in range(2, len(window_ids) + 2):  # Start from row 2
            cell = worksheet.cell(row=row, column=1)
            cell.fill = PatternFill(start_color='F2F2F2', fill_type='solid')
            
            # Replace window ID with actual date if available
            window_id = window_ids[row-2]  # Get the original window ID
            if window_id in window_dates:
                # Set the date value and format as date
                cell.value = window_dates[window_id]
                cell.number_format = 'yyyy-mm-dd'
            else:
                # Fallback to Window X format
                cell.value = f"Window {window_id}"
        
        # Rename the index column header
        worksheet.cell(row=1, column=1).value = "Date"
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 15  # Date column
        for col in range(2, len(all_factors) + 2):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 14
        
        # Add filters
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(all_factors) + 1)}1"
        
        # Create a summary sheet
        summary_data = [{
            'Number of Factors': len(all_factors),
            'Number of Windows': len(window_ids),
            'Total Predictions': df.count().sum(),
            'Average Prediction': df.mean().mean()
        }]
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        summary_sheet = writer.sheets['Summary']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col in range(1, 5):  # Columns A to D
            cell = summary_sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths for summary
        summary_sheet.column_dimensions['A'].width = 20  # Number of Factors
        summary_sheet.column_dimensions['B'].width = 20  # Number of Windows
        summary_sheet.column_dimensions['C'].width = 20  # Total Predictions
        summary_sheet.column_dimensions['D'].width = 20  # Average Prediction
        
        # Format numbers in summary
        for row in range(2, len(summary_data) + 2):
            # Format as integer
            for col in ['A', 'B', 'C']:
                cell = summary_sheet[f"{col}{row}"]
                cell.number_format = '0'
            # Format as decimal
            summary_sheet[f"D{row}"].number_format = '0.0000'


if __name__ == "__main__":
    logging.info("Starting XGBoost prediction generation...")
    try:
        main()
        logging.info("Prediction generation completed successfully.")
    except Exception as e:
        logging.error(f"Error in prediction generation: {str(e)}", exc_info=True)
        raise
